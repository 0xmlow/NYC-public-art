#!/usr/bin/env python3
"""
Build the unified Painted City dataset from multiple official sources.

Sources:
  1. NYC Parks Monuments (3762 historical + current works)
  2. NYC DOT Art Program (527 rotating street art commissions)
  3. Curated editorial picks (high-profile works with rich copy)

Outputs:
  - public/data/artworks.json   (used by the web map)
  - public/data/painted_city_dataset.xlsx (browsable spreadsheet)
"""

import csv
import json
import re
import os
import urllib.parse
from pathlib import Path
from pyproj import Transformer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "public" / "data"
OUT_DATA_DIR = ROOT / "public" / "data"
OUT_DATA_DIR.mkdir(parents=True, exist_ok=True)

# NYC State Plane Long Island (US feet) -> WGS84
SP_TO_WGS = Transformer.from_crs("EPSG:2263", "EPSG:4326", always_xy=True)

NYC_BBOX = (-74.30, 40.48, -73.68, 40.93)  # loose bounding box for sanity checks


def clean(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def titlecase_if_lower(s):
    if not s:
        return s
    if s.islower():
        return s.title()
    return s


def norm_borough(b):
    if not b:
        return "Unknown"
    b = b.strip().lower()
    if "manhattan" in b:
        return "Manhattan"
    if "bronx" in b:
        return "Bronx"
    if "brooklyn" in b:
        return "Brooklyn"
    if "queens" in b:
        return "Queens"
    if "staten" in b or "richmond" in b:
        return "Staten Island"
    if "all" in b:
        return "Citywide"
    return "Unknown"


def norm_type(t):
    if not t:
        return "Other"
    t = t.lower()
    if "mural" in t:
        return "Mural"
    if "sculpt" in t or "statue" in t or "bust" in t or "monument" in t:
        return "Sculpture"
    if "instal" in t or "interact" in t or "fiber" in t or "light" in t:
        return "Installation"
    if "fountain" in t:
        return "Fountain"
    if "plaque" in t or "tablet" in t or "inscrip" in t:
        return "Plaque"
    if "relief" in t:
        return "Relief"
    if "sign" in t:
        return "Signage"
    return "Other"


def in_nyc(lon, lat):
    try:
        lon = float(lon)
        lat = float(lat)
    except Exception:
        return False
    return NYC_BBOX[0] <= lon <= NYC_BBOX[2] and NYC_BBOX[1] <= lat <= NYC_BBOX[3]


def _google_search_url(title, artist):
    """Build a Google search URL for an artwork. Used as the primary
    'View source' target for non-curated entries — Google reliably
    surfaces Wikipedia / Parks / news / image results, where Wikipedia's
    strict full-text search would 0-result on obscure monuments."""
    parts = [(title or "").strip()]
    a = (artist or "").strip()
    if a and a.lower() != "unknown artist":
        parts.append(a)
    parts.append("NYC public art")
    q = " ".join(p for p in parts if p)
    return "https://www.google.com/search?q=" + urllib.parse.quote(q)


# ------------------------------------------------------------------
# 1. NYC Parks Monuments
# ------------------------------------------------------------------
def load_parks_monuments():
    items = []
    path = DATA_DIR / "nyc_parks_monuments.csv"
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            x = row.get("X")
            y = row.get("Y")
            if not x or not y:
                continue
            try:
                lon, lat = SP_TO_WGS.transform(float(x), float(y))
            except Exception:
                continue
            if not in_nyc(lon, lat):
                continue

            name = clean(row.get("name"))
            if not name:
                continue
            name = titlecase_if_lower(name)

            # Skip minimal/uninteresting plaques
            desc_parts = []
            if row.get("descrip"):
                desc_parts.append(clean(row["descrip"]))
            if row.get("Categories"):
                cats = clean(row["Categories"])
                if cats:
                    desc_parts.append(f"Category: {cats}")
            if row.get("materials"):
                desc_parts.append(f"Materials: {clean(row['materials'])}")
            if row.get("dimen"):
                desc_parts.append(f"Dimensions: {clean(row['dimen'])}")
            if row.get("sponsor"):
                desc_parts.append(f"Sponsor: {clean(row['sponsor'])}")
            if row.get("inscribed"):
                insc = clean(row["inscribed"])
                if insc:
                    desc_parts.append(f'Inscription: "{insc[:300]}"')
            description = " · ".join(p for p in desc_parts if p) or "NYC Parks monument."

            artist = ""
            for key in ("sculptor", "architect", "fabricator", "foundry"):
                if row.get(key):
                    artist = titlecase_if_lower(clean(row[key]))
                    break
            if not artist:
                artist = "Unknown artist"

            year = ""
            for key in ("dedicated", "cast", "Installation"):
                if row.get(key):
                    m = re.search(r"(1[6-9]\d{2}|20\d{2})", row[key])
                    if m:
                        year = m.group(1)
                        break

            # type — Parks monuments are usually sculptures / plaques
            category = (row.get("Categories") or "").lower()
            descrip = (row.get("descrip") or "").lower()
            combined = category + " " + descrip
            art_type = norm_type(combined) if combined.strip() else "Sculpture"

            location = clean(row.get("Location"))
            parkname = clean(row.get("parkname"))
            site = ", ".join(p for p in [parkname, location] if p) or "NYC Parks"

            extant = clean(row.get("extant")).upper()
            status = "Extant" if extant in ("Y", "YES", "TRUE", "1") else ("Removed" if extant in ("N", "NO") else "Unknown")

            # Build the official NYC Parks per-monument URL when both
            # the park number and monument number are present. This is
            # only useful as a fallback — the parks site currently
            # bot-blocks (403), but real-browser traffic *can* still
            # get through, and they may relax the block in the future.
            parknumber = clean(row.get("parknumber"))
            number = clean(row.get("number"))
            parks_link = ""
            if parknumber and number:
                parks_link = f"https://www.nycgovparks.org/parks/{parknumber}/monuments/{number}"

            items.append({
                "id": f"parks-{number or row.get('fileorder') or len(items)}",
                "source": "NYC Parks Monuments",
                "title": name,
                "artist": artist,
                "year": year,
                "borough": norm_borough(row.get("borough")),
                "type": art_type,
                "location": site,
                "lon": round(lon, 6),
                "lat": round(lat, 6),
                "description": description[:800],
                "materials": clean(row.get("materials")),
                "dimensions": clean(row.get("dimen")),
                "sponsor": clean(row.get("sponsor")),
                "donor": clean(row.get("donor")),
                "inscription": clean(row.get("inscribed"))[:500],
                "status": status,
                "image_url": "",
                # Primary "View source" target — Google search.
                # Google reliably surfaces the Wikipedia article (when one
                # exists), the NYC Parks page, news coverage, and image
                # results, so it's a more useful default than Wikipedia's
                # strict full-text search for obscure monuments. We drop
                # 'Unknown artist' from the query because it tanks recall.
                "source_link": _google_search_url(name, artist),
                # Secondary fallback — official NYC Parks per-monument page.
                "parks_link": parks_link,
            })
    return items


# ------------------------------------------------------------------
# 2. NYC DOT Art Program
# ------------------------------------------------------------------
def load_dot_art():
    items = []
    path = DATA_DIR / "nyc_dot_art.csv"
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            lat = row.get("Latitude")
            lon = row.get("Longitude")
            if not lat or not lon:
                continue
            try:
                lon_f = float(lon)
                lat_f = float(lat)
            except Exception:
                continue
            # DOT CSV has some bad longitudes. Filter those.
            if not in_nyc(lon_f, lat_f):
                continue

            title = clean(row.get("Title"))
            if not title:
                continue
            artist = clean(row.get("Artist")) or "Unknown artist"
            year = clean(row.get("Year"))
            borough = norm_borough(row.get("Borough"))
            project_type = clean(row.get("Project Type"))
            art_type = norm_type(project_type)
            site = clean(row.get("Site Location"))
            partner = clean(row.get("Partner"))
            program = clean(row.get("Program/Initiative"))
            inst_date = clean(row.get("Installation"))
            rem_date = clean(row.get("Removal"))

            desc_parts = []
            if program:
                desc_parts.append(f"Program: {program}")
            if partner:
                desc_parts.append(f"Partner: {partner}")
            if project_type:
                desc_parts.append(f"Medium: {project_type}")
            if inst_date:
                desc_parts.append(f"Installed: {inst_date}")
            if rem_date:
                desc_parts.append(f"Removed: {rem_date}")
            description = " · ".join(desc_parts) or "NYC DOT public art commission."

            items.append({
                "id": f"dot-{idx}",
                "source": "NYC DOT Art Program",
                "title": title,
                "artist": artist,
                "year": year,
                "borough": borough,
                "type": art_type,
                "location": site,
                "lon": round(lon_f, 6),
                "lat": round(lat_f, 6),
                "description": description[:800],
                "materials": project_type,
                "dimensions": "",
                "sponsor": partner,
                "donor": "",
                "inscription": "",
                "status": "Extant" if not rem_date else "Temporary / Removed",
                "image_url": "",
                "source_link": _google_search_url(title, artist),
            })
    return items


# ------------------------------------------------------------------
# 3. Curated editorial picks — high-profile works with rich copy
#    and verified Wikimedia images.
# ------------------------------------------------------------------
CURATED = [
    {
        "id": "curated-charging-bull",
        "title": "Charging Bull",
        "artist": "Arturo Di Modica",
        "year": "1989",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "Bowling Green, Lower Manhattan",
        "lon": -74.013382,
        "lat": 40.705568,
        "description": "A 7,100-pound bronze dropped outside the New York Stock Exchange overnight in December 1989 as an act of 'guerrilla art.' Di Modica funded it himself as a response to the 1987 stock market crash, intending it as a symbol of the strength and power of the American people. The city initially impounded it; public outcry forced its return to nearby Bowling Green, where it has become one of the most photographed sculptures in the world.",
        "artist_statement": "\"My point was to show people that if you want to do something in a moment things are very bad, you can do it. You can do it by yourself.\" \u2014 Arturo Di Modica",
        "materials": "Bronze",
        "dimensions": "11 ft tall, 16 ft long, 3.5 tons",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/2/23/Charging_Bull_at_Bowling_Green.jpg",
        "source_link": "https://en.wikipedia.org/wiki/Charging_Bull",
        "sponsor": "Self-funded by artist",
        "donor": "",
        "inscription": "",
        "status": "Extant",
    },
    {
        "id": "curated-fearless-girl",
        "title": "Fearless Girl",
        "artist": "Kristen Visbal",
        "year": "2017",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "Broad St & Wall St, Financial District",
        "lon": -74.011326,
        "lat": 40.706921,
        "description": "Installed on the eve of International Women's Day 2017, directly facing Di Modica's Charging Bull. Commissioned by State Street Global Advisors to promote its gender diversity index fund (ticker: SHE), she became an instant global icon of feminist resolve. Di Modica protested the installation as altering the meaning of his work; in 2018 Fearless Girl was relocated to face the NYSE itself.",
        "artist_statement": "\"She's there for all of us \u2014 she's there for women, she's there for girls, she's there for all the women of color and those who are marginalized.\" \u2014 Kristen Visbal",
        "materials": "Bronze",
        "dimensions": "4 ft tall, 250 lbs",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/e/e3/Fearless_Girl_statue_at_Bowling_Green.jpg/1024px-Fearless_Girl_statue_at_Bowling_Green.jpg",
        "source_link": "https://en.wikipedia.org/wiki/Fearless_Girl",
        "sponsor": "State Street Global Advisors",
        "donor": "",
        "inscription": "Know the power of women in leadership. SHE makes a difference.",
        "status": "Extant",
    },
    {
        "id": "curated-crack-is-wack",
        "title": "Crack is Wack",
        "artist": "Keith Haring",
        "year": "1986",
        "borough": "Manhattan",
        "type": "Mural",
        "location": "Crack is Wack Playground, 128th St & 2nd Ave, East Harlem",
        "lon": -73.932167,
        "lat": 40.802528,
        "description": "Haring painted this unauthorized handball court mural in 1986 as a direct response to the crack epidemic ravaging New York. He was arrested for the act; public pressure led the city to sanction and later restore the work. It now stands as one of Haring's most enduring public statements \u2014 an urgent, legible public service announcement rendered in his signature cartoon vocabulary. The site was officially landmarked by the NYC Parks Department.",
        "artist_statement": "\"I was driven to do something about the crack problem. It was something I had to do. And the fact that I got arrested for it only made it stronger.\" \u2014 Keith Haring",
        "materials": "Acrylic on concrete",
        "dimensions": "Full handball court (approx 20 ft tall)",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/1/17/Crack_is_Wack_mural_jeh.JPG/1024px-Crack_is_Wack_mural_jeh.JPG",
        "source_link": "https://www.nycgovparks.org/parks/crack-is-wack-playground",
        "sponsor": "Self-initiated; later adopted by NYC Parks",
        "donor": "",
        "inscription": "CRACK IS WACK",
        "status": "Extant (landmarked)",
    },
    {
        "id": "curated-vessel",
        "title": "Vessel",
        "artist": "Thomas Heatherwick",
        "year": "2019",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "Hudson Yards, 10th Ave at 30th St",
        "lon": -74.002192,
        "lat": 40.753710,
        "description": "A 150-foot climbable honeycomb of 154 interlocking staircases and 2,500 steps, clad in copper-colored steel. Inspired by Indian stepwells, Heatherwick conceived it as a 'vertical public square' \u2014 a structure that people could climb through rather than look at. It opened to acclaim in March 2019, was closed in 2021 after four suicides, and partially reopened in 2024 with safety netting.",
        "artist_statement": "\"I wanted to make something people could climb on, touch, explore \u2014 a landmark where visitors become part of the art.\" \u2014 Thomas Heatherwick",
        "materials": "Painted steel, polished copper-clad exterior",
        "dimensions": "150 ft tall, 80 ft wide at base, 150 ft at top",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/7/75/Hudson_Yards_Vessel_2019.jpg/1024px-Hudson_Yards_Vessel_2019.jpg",
        "source_link": "https://en.wikipedia.org/wiki/Vessel_(structure)",
        "sponsor": "Related Companies, Hudson Yards",
        "donor": "",
        "inscription": "",
        "status": "Extant (partially reopened)",
    },
    {
        "id": "curated-bowery-wall",
        "title": "The Houston Bowery Wall",
        "artist": "Rotating: Keith Haring, Os G\u00eameos, Shepard Fairey, Swoon, Banksy, Crash, and others",
        "year": "Since 1982",
        "borough": "Manhattan",
        "type": "Mural",
        "location": "E Houston St & Bowery, Lower East Side",
        "lon": -73.992893,
        "lat": 40.722535,
        "description": "The most storied mural wall in New York. Keith Haring painted it first in 1982 as a gift to the neighborhood; the wall sat fallow through the '90s until Goldman Global Arts revived it in 2008 as a curated rotating canvas. Since then it has been repainted by some of the most significant street artists alive \u2014 often drawing crowds on installation days. Each cycle runs a few months before being replaced.",
        "artist_statement": "\"The wall is a dialogue. Every artist who paints it is in conversation with Haring and with whoever comes next.\" \u2014 Jessica Goldman Srebnick, curator",
        "materials": "Acrylic and spray paint on concrete block",
        "dimensions": "Approx. 20 ft x 80 ft",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://en.wikipedia.org/wiki/Houston_Bowery_Wall",
        "sponsor": "Goldman Global Arts / Goldman Properties",
        "donor": "",
        "inscription": "",
        "status": "Extant (rotating)",
    },
    {
        "id": "curated-gay-liberation",
        "title": "Gay Liberation",
        "artist": "George Segal",
        "year": "1980",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "Christopher Park, Greenwich Village",
        "lon": -74.002136,
        "lat": 40.733700,
        "description": "Four life-size bronze figures \u2014 two standing men, two seated women \u2014 painted stark white in Segal's signature plaster-cast style. Installed in Christopher Park, directly across from the Stonewall Inn, as a tribute to the 1969 uprising that launched the modern LGBTQ+ rights movement. Controversial at the time; delayed for years before its unveiling. Now part of the Stonewall National Monument.",
        "artist_statement": "\"The sculpture must be loving and caring, and show the affection that is the hallmark of gay people.\" \u2014 George Segal",
        "materials": "Bronze with white lacquer finish",
        "dimensions": "Four life-size figures",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/d/d4/Gay_Liberation_Monument.JPG/1024px-Gay_Liberation_Monument.JPG",
        "source_link": "https://en.wikipedia.org/wiki/Gay_Liberation_Monument",
        "sponsor": "Mildred Andrews Fund / Peter Putnam",
        "donor": "",
        "inscription": "",
        "status": "Extant (National Monument)",
    },
    {
        "id": "curated-alamo",
        "title": "Alamo (The Cube)",
        "artist": "Tony Rosenthal",
        "year": "1967",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "Astor Place, East Village",
        "lon": -73.991160,
        "lat": 40.729838,
        "description": "A 15-foot black steel cube balanced on one corner at Astor Place, installed as part of the 'Sculpture in Environment' program in 1967. It was meant to be temporary; public affection kept it in place. Famously, it rotates on its axis \u2014 a tradition since the first students discovered you could push it. A meeting spot, skate backdrop, and unofficial symbol of the East Village.",
        "artist_statement": "\"I wanted to make a sculpture that the public could touch, that would become part of their daily lives.\" \u2014 Tony Rosenthal",
        "materials": "Cor-Ten steel, painted black",
        "dimensions": "15 ft cube",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/4/4a/Astor_Place_Cube_1.jpg/1024px-Astor_Place_Cube_1.jpg",
        "source_link": "https://en.wikipedia.org/wiki/Alamo_(Rosenthal)",
        "sponsor": "NYC Department of Cultural Affairs",
        "donor": "",
        "inscription": "",
        "status": "Extant",
    },
    {
        "id": "curated-brick-house",
        "title": "Brick House",
        "artist": "Simone Leigh",
        "year": "2019",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "High Line Plinth at 30th St, Hudson Yards",
        "lon": -74.004941,
        "lat": 40.752710,
        "description": "A 16-foot bronze bust of a Black woman whose torso fuses the architectural form of a clay house \u2014 referencing the dwellings of the Mousgoum people of Cameroon, the Batammaliba of Togo and Benin, and restaurant architecture of the American South. Leigh's monumental figure inaugurated the High Line Plinth, claiming the Hudson Yards skyline for Black womanhood.",
        "artist_statement": "\"Brick House is a physical embodiment of Black women \u2014 their endurance, their architecture, their presence.\" \u2014 Simone Leigh",
        "materials": "Bronze",
        "dimensions": "16 ft tall",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://www.thehighline.org/art/projects/simone-leigh/",
        "sponsor": "High Line Art",
        "donor": "",
        "inscription": "",
        "status": "Extant",
    },
    {
        "id": "curated-dinosaur",
        "title": "Dinosaur",
        "artist": "Iv\u00e1n Argote",
        "year": "2024",
        "borough": "Manhattan",
        "type": "Sculpture",
        "location": "High Line Plinth at 30th St, Hudson Yards",
        "lon": -74.004850,
        "lat": 40.752810,
        "description": "A hyper-realistic 16-foot aluminum pigeon perched on a pedestal at the High Line Plinth \u2014 an ironic monument to the unglamorous 'immigrant' bird that actually dominates New York. Argote's piece pokes at the tradition of bronze heroes on pedestals while celebrating the commonest creature in the city.",
        "artist_statement": "\"Pigeons are the true New Yorkers. They are travelers, they adapt, they survive. They are us.\" \u2014 Iv\u00e1n Argote",
        "materials": "Painted aluminum",
        "dimensions": "16 ft tall",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://www.thehighline.org/art/projects/ivan-argote/",
        "sponsor": "High Line Art",
        "donor": "",
        "inscription": "",
        "status": "Extant",
    },
    {
        "id": "curated-audrey",
        "title": "Audrey of Mulberry",
        "artist": "Tristan Eaton",
        "year": "2013",
        "borough": "Manhattan",
        "type": "Mural",
        "location": "176 Mulberry Street, Little Italy",
        "lon": -73.996500,
        "lat": 40.721000,
        "description": "A layered pop-art portrait of Audrey Hepburn on the side of 176 Mulberry Street, commissioned by the LISA Project NYC. Eaton's signature collage style fills Hepburn's face with dozens of vintage illustrations, ads, and pop-culture fragments. It has become one of Little Italy's most-photographed murals.",
        "artist_statement": "\"Audrey was an icon of grace. I wanted to fill her with the chaos of New York.\" \u2014 Tristan Eaton",
        "materials": "Acrylic and spray paint on brick",
        "dimensions": "Approx. 30 ft tall",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://thelisaprojectnyc.org/",
        "sponsor": "The LISA Project NYC",
        "donor": "",
        "inscription": "",
        "status": "Extant",
    },
    {
        "id": "curated-big-pun",
        "title": "Big Pun Memorial",
        "artist": "Tats Cru",
        "year": "2000 (repainted annually)",
        "borough": "Bronx",
        "type": "Mural",
        "location": "910 Rogers Place, Longwood, South Bronx",
        "lon": -73.902500,
        "lat": 40.827500,
        "description": "The legendary memorial to Christopher Rios \u2014 Big Pun \u2014 the first solo Latino rapper to go platinum, who died in 2000 at age 28. Tats Cru, the Bronx collective that rose from 1980s graffiti to become the borough's visual ambassadors, repaints the mural every year on his birthday, a ritual that has made it one of the most documented street art sites in New York.",
        "artist_statement": "\"Big Pun belonged to the block. We keep him here.\" \u2014 BG183 of Tats Cru",
        "materials": "Aerosol on brick",
        "dimensions": "Full building wall",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://citylore.org/places/big-pun-mural/",
        "sponsor": "Community / Tats Cru",
        "donor": "",
        "inscription": "Big Pun 1971-2000",
        "status": "Extant (repainted annually)",
    },
    {
        "id": "curated-welling-court",
        "title": "Welling Court Mural Project",
        "artist": "150+ rotating artists (curated by Ad Hoc Art)",
        "year": "Since 2009",
        "borough": "Queens",
        "type": "Mural",
        "location": "Welling Court, Astoria",
        "lon": -73.922500,
        "lat": 40.772500,
        "description": "A residential block in Astoria turned into a rotating outdoor gallery of 150+ murals, refreshed annually each June. Founded in 2009 by Georgia and Jonathan Lee (Ad Hoc Art) at the invitation of a local resident, it's now one of the largest and longest-running mural projects in NYC \u2014 painted on garage doors, apartment walls, and private homes. Unlike curated outdoor museums, the works here live alongside actual residents.",
        "artist_statement": "\"This is not a gallery. These are people's homes. Every mural is a gift and a conversation.\" \u2014 Georgia Lee, co-founder",
        "materials": "Aerosol, acrylic, mixed media",
        "dimensions": "Full neighborhood (multiple blocks)",
        "source": "Curated",
        "image_url": "",
        "source_link": "https://www.wellingcourtmuralproject.com/",
        "sponsor": "Ad Hoc Art",
        "donor": "",
        "inscription": "",
        "status": "Extant (rotating annually)",
    },
    {
        "id": "curated-unisphere",
        "title": "Unisphere",
        "artist": "Gilmore D. Clarke",
        "year": "1964",
        "borough": "Queens",
        "type": "Sculpture",
        "location": "Flushing Meadows Corona Park",
        "lon": -73.845200,
        "lat": 40.746050,
        "description": "A 140-foot-tall stainless-steel globe commissioned by U.S. Steel for the 1964 World's Fair, themed 'Peace Through Understanding.' It remains the largest spherical structure in the world and has been designated an official NYC landmark. The three orbiting rings symbolize the first satellites: Yuri Gagarin, John Glenn, and Telstar.",
        "artist_statement": "",
        "materials": "Stainless steel",
        "dimensions": "140 ft tall, 120 ft diameter, 700,000 lbs",
        "source": "Curated",
        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/8/8a/Unisphere_3.jpg/1024px-Unisphere_3.jpg",
        "source_link": "https://en.wikipedia.org/wiki/Unisphere",
        "sponsor": "U.S. Steel",
        "donor": "",
        "inscription": "",
        "status": "Extant (NYC Landmark)",
    },
]


# ------------------------------------------------------------------
# Merge + deduplicate
# ------------------------------------------------------------------
def load_community_additions():
    """Merge in community-submitted entries from data/community_additions.json.

    Maintainers add entries here after reviewing issues filed via the
    'Submit an artwork' GitHub Issue template. Each record follows the same
    shape as CURATED entries (title, artist, lon, lat, type, borough, ...).
    """
    path = DATA_DIR / "community_additions.json"
    if not path.exists():
        return []
    try:
        raw = json.loads(path.read_text())
    except Exception as e:
        print(f"  ! could not parse community_additions.json: {e}")
        return []
    out = []
    for i, rec in enumerate(raw):
        # Sanity: must have title + coords
        if not rec.get("title") or rec.get("lon") is None or rec.get("lat") is None:
            continue
        rec = dict(rec)  # shallow copy
        rec.setdefault("id", f"community-{i}")
        rec.setdefault("source", "Community")
        rec.setdefault("borough", "")
        rec.setdefault("type", "Other")
        out.append(rec)
    return out


def merge_all():
    parks = load_parks_monuments()
    dot = load_dot_art()
    community = load_community_additions()
    print(f"Parks monuments: {len(parks)}")
    print(f"DOT art: {len(dot)}")
    print(f"Curated: {len(CURATED)}")
    print(f"Community: {len(community)}")

    # Deduplicate curated against parks by title match
    curated_titles = {c["title"].lower() for c in CURATED}
    parks_filtered = [p for p in parks if p["title"].lower() not in curated_titles]

    all_items = CURATED + community + parks_filtered + dot

    # Dedupe only exact duplicates (same title at same exact 6-decimal coord)
    seen = set()
    unique = []
    for it in all_items:
        key = (it["title"].lower().strip(), it["lon"], it["lat"], it.get("artist", "").lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append(it)

    # Sort: curated first, then community, then parks, then dot
    def sort_key(it):
        src = it.get("source", "")
        if src == "Curated":
            return (0, it["title"])
        if src == "Community":
            return (1, it["title"])
        if "Parks" in src:
            return (2, it["title"])
        return (3, it["title"])

    unique.sort(key=sort_key)
    print(f"Total unique: {len(unique)}")
    return unique


# ------------------------------------------------------------------
# Write JSON
# ------------------------------------------------------------------
def write_json(items):
    out = OUT_DATA_DIR / "artworks.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=1)
    print(f"Wrote {out} ({out.stat().st_size // 1024} KB)")


# ------------------------------------------------------------------
# Write XLSX
# ------------------------------------------------------------------
def write_xlsx(items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Painted City"

    columns = [
        ("id", 22),
        ("title", 36),
        ("artist", 32),
        ("year", 10),
        ("borough", 14),
        ("type", 14),
        ("location", 40),
        ("lat", 12),
        ("lon", 12),
        ("description", 60),
        ("artist_statement", 60),
        ("materials", 26),
        ("dimensions", 26),
        ("sponsor", 26),
        ("donor", 20),
        ("inscription", 40),
        ("status", 18),
        ("image_url", 40),
        ("source", 22),
        ("source_link", 40),
    ]

    header_fill = PatternFill("solid", fgColor="14141B")
    header_font = Font(color="FFB84D", bold=True, name="Helvetica")
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(items, 2):
        for col_idx, (name, _) in enumerate(columns, 1):
            val = item.get(name, "") if name != "artist_statement" else item.get("artist_statement", "")
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = OUT_DATA_DIR / "painted_city_dataset.xlsx"
    wb.save(out)
    print(f"Wrote {out} ({out.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    items = merge_all()
    write_json(items)
    write_xlsx(items)
    print(f"\nDone. {len(items)} artworks in dataset.")
